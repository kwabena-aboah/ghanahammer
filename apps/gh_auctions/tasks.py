"""
GhanaHammer Celery Tasks
Auction lifecycle, auto-bid processing, notifications, bulk import, analytics
"""
import logging
from celery import shared_task
from django.utils import timezone
from django.conf import settings

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# AUCTION LIFECYCLE
# ─────────────────────────────────────────────────────────────────────────────

@shared_task(name='close_expired_auctions')
def close_expired_auctions():
    """
    Run every 60 seconds.
    Close auctions that have passed their end_time.
    Determine winner, update status, trigger notifications.
    """
    from apps.gh_auctions.models import Auction
    from apps.bidding.models import Bid
    from apps.notifications.services import NotificationService

    now = timezone.now()
    expired = Auction.objects.filter(
        status__in=['active', 'extended', 'closing'],
        end_time__lte=now,
    ).select_related('seller')

    closed_count = 0
    for auction in expired:
        try:
            winning_bid = Bid.objects.filter(
                auction=auction, status=Bid.STATUS_WINNING
            ).select_related('bidder').first()

            if winning_bid and (not auction.reserve_price or winning_bid.amount >= auction.reserve_price):
                # Auction SOLD
                auction.status = Auction.STATUS_SOLD
                auction.winner = winning_bid.bidder
                auction.winning_bid_amount = winning_bid.amount
                auction.save(update_fields=['status', 'winner', 'winning_bid_amount'])

                winning_bid.status = Bid.STATUS_WON
                winning_bid.save(update_fields=['status'])

                # Update winner stats
                winning_bid.bidder.total_wins += 1
                winning_bid.bidder.save(update_fields=['total_wins'])

                # Notify winner and seller
                NotificationService.notify_auction_won(winning_bid.bidder, auction, winning_bid.amount)
                NotificationService.notify_seller_sale(auction.seller, auction, winning_bid.amount)

                logger.info(f'Auction {auction.lot_number} SOLD to {winning_bid.bidder.email} for GHS {winning_bid.amount}')
            else:
                # No winning bid or reserve not met
                auction.status = Auction.STATUS_UNSOLD
                auction.save(update_fields=['status'])

                # Notify seller
                NotificationService.create_in_app(
                    auction.seller,
                    f'Auction ended unsold — {auction.lot_number}',
                    f'"{auction.title}" ended without meeting the reserve price. Consider re-listing.',
                    'warning',
                    auction,
                )
                logger.info(f'Auction {auction.lot_number} ended UNSOLD')

            closed_count += 1
        except Exception as e:
            logger.exception(f'Error closing auction {auction.lot_number}: {e}')

    return f'Closed {closed_count} auctions'


@shared_task(name='activate_scheduled_auctions')
def activate_scheduled_auctions():
    """Activate auctions that have reached their start_time"""
    from apps.gh_auctions.models import Auction
    from apps.notifications.services import NotificationService
    from apps.gh_accounts.models import SavedSearch

    now = timezone.now()
    to_activate = Auction.objects.filter(
        status='scheduled',
        start_time__lte=now,
    )

    for auction in to_activate:
        # Check if VIP preview phase is still active
        if auction.preview_start and auction.preview_start > now:
            continue  # Not yet time

        auction.status = Auction.STATUS_ACTIVE
        auction.save(update_fields=['status'])
        logger.info(f'Auction {auction.lot_number} activated')

        # Notify saved search watchers
        check_saved_searches_for_auction.delay(str(auction.id))

    return f'Activated {to_activate.count()} auctions'


@shared_task(name='activate_vip_previews')
def activate_vip_previews():
    """Move auctions from scheduled to preview status"""
    from apps.gh_auctions.models import Auction

    now = timezone.now()
    Auction.objects.filter(
        status='scheduled',
        preview_start__lte=now,
        start_time__gt=now,
    ).update(status=Auction.STATUS_PREVIEW)


@shared_task(name='check_saved_searches_for_auction')
def check_saved_searches_for_auction(auction_id):
    """Match a new auction against all active saved searches"""
    from apps.gh_auctions.models import Auction
    from apps.gh_accounts.models import SavedSearch
    from apps.notifications.services import NotificationService
    from django.db.models import Q

    try:
        auction = Auction.objects.select_related('category').get(id=auction_id)
    except Auction.DoesNotExist:
        return

    # Find matching saved searches
    searches = SavedSearch.objects.filter(
        is_active=True,
        alert_new_items=True,
    ).select_related('user')

    for search in searches:
        matched = True

        if search.query and search.query.lower() not in auction.title.lower():
            matched = False
        if search.category and search.category.lower() not in auction.category.name.lower():
            matched = False
        if search.min_price and auction.starting_price < search.min_price:
            matched = False
        if search.max_price and auction.starting_price > search.max_price:
            matched = False
        if search.location and search.location.lower() not in (auction.location or '').lower():
            matched = False
        if search.condition and search.condition != auction.condition:
            matched = False

        if matched:
            NotificationService.notify_new_matching_item(search.user, auction, search)


# ─────────────────────────────────────────────────────────────────────────────
# BULK IMPORT
# ─────────────────────────────────────────────────────────────────────────────

@shared_task(name='process_bulk_import')
def process_bulk_import(batch_id):
    """Process a CSV or Excel bulk import batch"""
    from apps.gh_auctions.models import BulkImportBatch, Auction, Category
    from django.utils.dateparse import parse_datetime
    from decimal import Decimal, InvalidOperation
    import csv
    import io

    try:
        batch = BulkImportBatch.objects.get(id=batch_id)
    except BulkImportBatch.DoesNotExist:
        return

    batch.status = 'processing'
    batch.save(update_fields=['status'])

    errors = []
    imported = 0
    failed = 0

    try:
        file_path = batch.file.path
        ext = file_path.lower().split('.')[-1]

        if ext == 'csv':
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        elif ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            batch.status = 'failed'
            batch.error_log = 'Unsupported file format'
            batch.save()
            return

        batch.total_rows = len(rows)
        batch.save(update_fields=['total_rows'])

        for i, row in enumerate(rows, start=2):
            try:
                title = str(row.get('title', '') or '').strip()
                if not title:
                    errors.append(f'Row {i}: Missing title')
                    failed += 1
                    continue

                cat_name = str(row.get('category', '') or '').strip()
                category = Category.objects.filter(
                    name__icontains=cat_name
                ).first() if cat_name else Category.objects.first()

                if not category:
                    errors.append(f'Row {i}: Category not found: {cat_name}')
                    failed += 1
                    continue

                starting_price = Decimal(str(row.get('starting_price', '10') or '10'))
                end_time_str = str(row.get('end_time', '') or '')
                end_time = parse_datetime(end_time_str)
                if not end_time:
                    from datetime import timedelta
                    end_time = timezone.now() + timedelta(days=7)

                start_time_str = str(row.get('start_time', '') or '')
                start_time = parse_datetime(start_time_str) or timezone.now()

                Auction.objects.create(
                    seller=batch.seller,
                    title=title,
                    category=category,
                    description=str(row.get('description', '') or ''),
                    condition=str(row.get('condition', 'used') or 'used'),
                    starting_price=starting_price,
                    current_price=starting_price,
                    reserve_price=Decimal(str(row['reserve_price'])) if row.get('reserve_price') else None,
                    buy_now_price=Decimal(str(row['buy_now_price'])) if row.get('buy_now_price') else None,
                    bid_increment=Decimal(str(row.get('bid_increment', '10') or '10')),
                    start_time=start_time,
                    end_time=end_time,
                    location=str(row.get('location', '') or ''),
                    region=str(row.get('region', '') or ''),
                    status='scheduled' if start_time > timezone.now() else 'active',
                    bulk_import_batch=str(batch_id),
                )
                imported += 1

            except (InvalidOperation, ValueError, Exception) as e:
                errors.append(f'Row {i}: {str(e)}')
                failed += 1

    except Exception as e:
        logger.exception(f'Bulk import error: {e}')
        batch.status = 'failed'
        batch.error_log = str(e)
        batch.save()
        return

    batch.imported = imported
    batch.failed = failed
    batch.error_log = '\n'.join(errors[:100])
    batch.status = 'done'
    batch.save()

    # Notify seller
    from apps.notifications.services import NotificationService
    NotificationService.create_in_app(
        batch.seller,
        f'Bulk import complete — {imported} listings created',
        f'{imported} auctions imported, {failed} failed. Check dashboard for details.',
        'success' if not failed else 'warning',
    )

    logger.info(f'Bulk import batch {batch_id}: {imported} imported, {failed} failed')
    return f'{imported} imported, {failed} failed'
